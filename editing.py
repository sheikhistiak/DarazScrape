import requests
from PIL import Image
from openpyxl import load_workbook, Workbook  # Import Workbook class
from io import BytesIO
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
# Sample data (replace with your actual data)
product_title = "Product Title"
product_price = "$19.99"
product_review_summary = "4.5/5"
product_promotion = "Limited Time Offer"
product_badge = "Daraz Mall Badge Found"
product_detail_title = "Product Details"
product_highlights = "Product Highlights"

# Download product image
product_image_url = "https://example.com/product_image.jpg"  # Replace with the actual URL
response = requests.get(product_image_url)
products_image = Image.open(BytesIO(response.content))

# Download badge image (if available)
badge_image_url = "https://example.com/badge_image.png"  # Replace with the actual URL
response = requests.get(badge_image_url)
badges_image = Image.open(BytesIO(response.content)) if response.status_code == 200 else None

# Load an existing workbook or create a new one
try:
    wb = load_workbook("product_data.xlsx")
except FileNotFoundError:
    wb = Workbook()

# Select the active sheet or create one if it doesn't exist
ws = wb.active if "Sheet" in wb.sheetnames else wb.create_sheet("Sheet")

# Append the data to the worksheet
row = [
    product_title,
    product_price,
    product_review_summary,
    product_promotion,
    product_badge,
    product_detail_title,
    product_highlights,
]

# Append None for image columns if badge image is not available
if badges_image is None:
    row.append(None)
    row.append(None)
else:
    row.append("Product Image")
    row.append("Badge Image")

ws.append(row)  # Append data to the worksheet

# Add product image
if products_image:
    ws.add_image(products_image, f"G{ws.max_row}")  # Insert the product image in column G

# Add badge image (if available)
if badges_image:
    ws.add_image(badges_image, f"H{ws.max_row}")  # Insert the badge image in column H

# Save the workbook
wb.save("product_data.xlsx")



# Define the image URL
image_url = product_image_url.split("_.webp")[0]  # Replace with the actual image URL
print(image_url)
# Download the image
response = requests.get(image_url)
if response.status_code == 200:
    image_bytes = BytesIO(response.content)
else:
    image_bytes = None

# Load the existing Excel workbook or create a new one
try:
    wb = load_workbook('excel_with_images.xlsx')
except FileNotFoundError:
    wb = Workbook()

# Select the desired worksheet (change 'Sheet' to the sheet name you want to work with)
ws = wb['Sheet']

# Determine the next available row in column A
next_row = len(ws['H']) + 1

# Insert the image into the worksheet
if image_bytes:
    image = Image.open(image_bytes)
    ws.add_image(ExcelImage(image),
                 f'A{next_row}')  # Insert the image in column A of the next available row

# Save the modified Excel workbook
wb.save('excel_with_images.xlsx')